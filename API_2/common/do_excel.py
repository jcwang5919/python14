from openpyxl import load_workbook
from API_2.common import project_path
from API_2.common.read_config import ReadConfig



class DoExcel:
    """该类完成测试数据的读取以及测试结果的写回"""

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self):
        """从Excel读取数据，有返回值"""
        case_id = ReadConfig(project_path.conf_path).get_data('CASE', 'case_id')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            for j in range(1, sheet.max_column + 1):
                row_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            test_data.append(row_data)
        wb.close()
        final_data = []  # 空列表 存储最终的测试用例数据
        if case_id == 'all':  # 如果case_id==all 那就获取所有的用例数据
            final_data = test_data  # 把测试用例赋值给final_data这个变量
        else:  # 否则 如果是列表 那就获取列表里面指定id的用例的数据
            for i in case_id:  # 遍历case_id 里面的值
                final_data.append(test_data[i - 1])
        return final_data

    def write_back(self, row, col, value):
        """写回测试结果到Excel中"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    test_data = DoExcel(project_path.case_path, 'test_case').read_data()
    print(test_data)
    print('字典的数据长度为：', len(test_data))
