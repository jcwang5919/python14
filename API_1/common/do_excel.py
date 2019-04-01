from openpyxl import load_workbook
from API_1.common import project_path


class DoExcel:
    """该类完成测试数据的读取以及测试结果的写回"""

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name  # Excel工作簿文件名或地址
        self.sheet_name = sheet_name  # 表单名

    def read_data(self):
        """从Excel读取数据，有返回值"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        data = []
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            for j in range(1, sheet.max_column + 1):
                row_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data.append(row_data)
        wb.close()
        return data

    def write_back(self, row, col, value):
        """写回测试结果到Excel中"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    test_data = DoExcel(project_path.case_path, 'test_case').read_data()
    print(test_data)
    print('字典的数据长度为：', len(test_data))
